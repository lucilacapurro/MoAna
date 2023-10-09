#Importo librerías
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy
from scipy import signal
from openpyxl import load_workbook
import os
from numpy.core.function_base import linspace
from ProcesamientoOnline import funcLevantarSenal, funcObtenerHR, funcFiltroBP, funcFiltroLP, funcFiltroHP, funcFiltrar, funcProcesamiento, funcDetectarPicos, funcEliminarPicosSubida, funcEliminarDiastolicos, funcDetectarOnsets, funcHBI, funcPPGA


############################################################################################################################################

fs = 25

# CAMBIAR CSV DATA
nombre_csv = 'DATA_XXX_25.csv'

directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_csv_senal_ppg = nombre_csv
path_csv_ppg = os.path.join(directorio_actual, nombre_csv_senal_ppg)
senal_ppg = funcLevantarSenal(path_csv_ppg)

# CAMBIAR MUESTRAS
senal_ppg = senal_ppg[:]


HR = funcObtenerHR(senal_ppg=senal_ppg, f1=0.1, f2=5, fs=fs)
senal_ppg_filtrada = funcFiltrar(senal_ppg, fs, HR, orden=3)
PPGA, HBI = funcProcesamiento(senal_ppg_filtrada, fs, HR)


# Excel Curva Normalizacion
directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_normalizacion = 'Curva Normalizacion.xlsx'
path_normalizacion = os.path.join(directorio_actual, nombre_excel_normalizacion)
excel_normalizacion = load_workbook(path_normalizacion)
hoja_excel_normalizacion = excel_normalizacion['Hoja1']

for i in range(len(HBI)):
    nueva_medicion = [HBI[i], PPGA[i], nombre_csv]
    hoja_excel_normalizacion.append(nueva_medicion)

excel_normalizacion.save(path_normalizacion)


# Excel Desvios 
directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_desvios = 'Desvios.xlsx'
path_desvios = os.path.join(directorio_actual, nombre_excel_desvios)
excel_desvios = load_workbook(path_desvios)
hoja_excel_desvios = excel_desvios['Hoja1']

desvio_HBI = np.std(HBI)
desvio_PPGA = np.std(PPGA)

desvios = [desvio_HBI, desvio_PPGA, nombre_csv]

hoja_excel_desvios.append(nueva_medicion)

excel_desvios.save(path_desvios)
