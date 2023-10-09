'''

ATENCION: OJO CAMBIAR LOS MIN Y MAX AL AGREGAR NUEVOS HBI Y PPGA A LA CURVA !!!

'''

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import signal
import os
from openpyxl import load_workbook

def funcHistograma(parametro, tipo_parametro):
  if tipo_parametro == "HBI":
    min = 0
    max = 100
  elif tipo_parametro == "PPGA":
    min = 0
    max = 100000
  bins = max-min
  vector_parametro = np.linspace(min, max, bins+1, dtype = int)
  vector_hist = np.zeros(bins+1)
  for i in range(len(parametro)):
    vector_hist[int(parametro[i])-min] += 1
  vector_hist_norm = vector_hist/np.sum(vector_hist)
  plt.figure(figsize=(20,5))
  plt.title('Histograma', fontsize=15)
  plt.bar(x = vector_parametro, height = vector_hist_norm, width=0.8, bottom=None, align='center')
  plt.show()
  TF_histograma = [[int(param), param_norm] for param, param_norm in zip(vector_parametro, vector_hist_norm)]
  return TF_histograma


def funcHistogramaAcumulado(TF_histograma):
  min = TF_histograma[0][0]
  max = TF_histograma[-1][0]
  bins = len(TF_histograma)
  vector_parametro = np.linspace(min, max, bins, dtype = int)
  vector_hist_acum = np.zeros(len(TF_histograma))
  vector_hist_acum[0] = TF_histograma[0][1]
  for i in range(1,len(TF_histograma)):
    vector_hist_acum[i] = vector_hist_acum[i-1] + TF_histograma[i][1]
  plt.figure(figsize=(10,5))
  plt.title('Histograma Acumulado', fontsize=15)
  plt.bar(x = vector_parametro, height = vector_hist_acum, width=0.8, bottom=None, align='center')
  plt.show()
  TF_histograma_acum = [[int(param), round(param_norm, 7)] for param, param_norm in zip(vector_parametro, vector_hist_acum)]
  return TF_histograma_acum


def funcNormalizarParametro(TF_normalizacion, parametro):
  parametro = int(parametro)
  min = TF_normalizacion[0][0]
  parametro_norm = TF_normalizacion[parametro-min][1]
  return parametro_norm


def funcPromedio(vector_valores):
  prom = np.mean(vector_valores)
  return int(prom)

def funcSPIi(PPGAi_norm, HBIi_norm):
  SPIi = 100 - (0.67*PPGAi_norm+0.33*HBIi_norm)*100
  return int(SPIi)


directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_normalizacion = 'Curva Normalizacion.xlsx'
path_normalizacion = os.path.join(directorio_actual, nombre_excel_normalizacion)

excel_normalizacion = pd.read_excel(path_normalizacion)

HBI_curva_excel = excel_normalizacion['HBI']
PPGA_curva_excel = excel_normalizacion['PPGA']

histograma_HBI = funcHistograma(HBI_curva_excel, "HBI")
TF_HBI = funcHistogramaAcumulado(histograma_HBI)

histograma_PPGA = funcHistograma(PPGA_curva_excel, "PPGA")
TF_PPGA = funcHistogramaAcumulado(histograma_PPGA)


# Guardamos las TF en los excels:
# HBI
nombre_excel_TF_HBI = 'TF_HBI.xlsx'
path_excel_TF_HBI = os.path.join(directorio_actual, nombre_excel_TF_HBI)
excel_TF_HBI = load_workbook(path_excel_TF_HBI)
df_TF_HBI = pd.DataFrame(TF_HBI)
df_TF_HBI.to_excel(path_excel_TF_HBI, index=False, header=False)

# PPGA
nombre_excel_TF_PPGA = 'TF_PPGA.xlsx'
path_excel_TF_PPGA = os.path.join(directorio_actual, nombre_excel_TF_PPGA)
excel_TF_PPGA = load_workbook(path_excel_TF_PPGA)
df_TF_PPGA = pd.DataFrame(TF_PPGA)
df_TF_PPGA.to_excel(path_excel_TF_PPGA, index=False, header=False)


'''
vectorSPI = []

for i in range(len(PPGA)): 
  PPGA_norm = funcNormalizarParametro(TF_PPGA, int(PPGA[i]))
  HBI_norm = funcNormalizarParametro(TF_HBI, int(HBI[i]))
  SPI = funcSPIi(PPGA_norm, HBI_norm)
  vectorSPI.append(SPI)

print(vectorSPI)


# Crear una lista de índices para el eje x
x = range(len(vectorSPI))

# Crear la figura y el gráfico de líneas
plt.figure()
plt.plot(x, vectorSPI)

# Configurar los títulos y etiquetas de los ejes
plt.title("")
plt.xlabel("")
plt.ylabel("HBI norm")
plt.ylim(0,100)

# Mostrar el gráfico
plt.show()
'''