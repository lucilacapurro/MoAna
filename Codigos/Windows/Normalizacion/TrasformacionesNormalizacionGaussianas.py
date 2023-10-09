import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import signal
import os
from openpyxl import load_workbook

def funcNormalizarParametro(TF_normalizacion, parametro):
  parametro = int(parametro)
  min = TF_normalizacion[0][0]
  parametro_norm = TF_normalizacion[parametro-min][1]
  return parametro_norm

def funcPromedio(vector_valores):
  prom = np.mean(vector_valores)
  return int(prom)

def funcSPIi(PPGAi_norm, HBIi_norm):
  SPIi = 100 - (0.67*PPGAi_norm+0.33*HBIi_norm)*100
  return int(SPIi)


# Obtenemos los valores de HBI y PPGA de entrada:
directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_normalizacion = 'Curva Normalizacion.xlsx'
path_normalizacion = os.path.join(directorio_actual, nombre_excel_normalizacion)
excel_normalizacion = pd.read_excel(path_normalizacion)

lista_poblacional_HBI = excel_normalizacion['HBI'].dropna()
lista_poblacional_PPGA = excel_normalizacion['PPGA'].dropna()

min_poblacional_HBI = np.min(lista_poblacional_HBI)
max_poblacional_HBI = np.max(lista_poblacional_HBI)
x_HBI = np.linspace(min_poblacional_HBI, max_poblacional_HBI, max_poblacional_HBI-min_poblacional_HBI+1)
x_HBI = x_HBI.tolist()

min_poblacional_PPGA = np.min(lista_poblacional_PPGA)
max_poblacional_PPGA = np.max(lista_poblacional_PPGA)
x_PPGA = np.linspace(min_poblacional_PPGA, max_poblacional_PPGA, max_poblacional_PPGA-min_poblacional_PPGA+1)
x_PPGA = x_PPGA.tolist()

# Obtenemos los valores de HBInorm y PPGAnorm de salida:
directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_normalizacion_gaussiana = 'Curva Normalizacion Gaussiana.xlsx'
path_normalizacion_gaussiana = os.path.join(directorio_actual, nombre_excel_normalizacion_gaussiana)
excel_normalizacion_gaussiana = pd.read_excel(path_normalizacion_gaussiana)

HBI_curva_excel = np.array(excel_normalizacion_gaussiana['HBI'].dropna())
PPGA_curva_excel = np.array(excel_normalizacion_gaussiana['PPGA'].dropna())

y_HBI = np.cumsum(HBI_curva_excel)
y_HBI = y_HBI.tolist()
y_PPGA = np.cumsum(PPGA_curva_excel)
y_PPGA = y_PPGA.tolist()

df_TF_HBI = pd.DataFrame({'x HBI': x_HBI, 'y HBI': y_HBI})
df_TF_PPGA = pd.DataFrame({'x PPGA': x_PPGA, 'y PPGA': y_PPGA})


# Guardamos las TF en los excels:
# HBI
nombre_excel_TF_HBI = 'TF_HBI_Gaussiana.xlsx'
path_excel_TF_HBI = os.path.join(directorio_actual, nombre_excel_TF_HBI)
excel_TF_HBI = load_workbook(path_excel_TF_HBI)
df_TF_HBI.to_excel(path_excel_TF_HBI, index=False, header=False)

# PPGA
nombre_excel_TF_PPGA = 'TF_PPGA_Gaussiana.xlsx'
path_excel_TF_PPGA = os.path.join(directorio_actual, nombre_excel_TF_PPGA)
excel_TF_PPGA = load_workbook(path_excel_TF_PPGA)
df_TF_PPGA.to_excel(path_excel_TF_PPGA, index=False, header=False)


############################################################################################

plt.figure(figsize=(10,5))
plt.title('Histograma Acumulado', fontsize=15)
plt.bar(x = x_HBI, height = y_HBI, width=0.8, bottom=None, align='center')
plt.show()

plt.figure(figsize=(10,5))
plt.title('Histograma Acumulado', fontsize=15)
plt.bar(x = x_PPGA, height = y_PPGA, width=0.8, bottom=None, align='center')
plt.show()
