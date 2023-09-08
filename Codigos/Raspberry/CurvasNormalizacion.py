import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy
from scipy import signal
from openpyxl import load_workbook
import os
from numpy.core.function_base import linspace

############################################################################################################################################
# FUNCIONES

# Función para obtener la señal de PPG a partir de los valores de R e IR 
def funcLevantarSenal(path_ppg):
  df_ppg = pd.read_csv(path_ppg)
  senal_ppg = df_ppg['PPG']
  return senal_ppg

# Función para visualizar la señal de PPG 
def funcVisualizar(senal, fs, label, scatterX = 0, scatterY = 0, scatter = False): # scatterX está en muestras --> genera vector tiempos = muestras/fs 
  L = len(senal)
  t = np.linspace(0,L/fs,L)
  plt.figure(figsize=(25,5))
  plt.plot(t,senal, 'y')
  if (scatter == True): 
    plt.scatter(scatterX/fs, scatterY)
  plt.ylabel('Amplitud', fontsize=14)
  plt.xlabel('Tiempo [s]', fontsize=14)
  plt.title(label, fontsize=14)
  plt.show()

# Función para filtrar la señal con un filtro pasa bajos 
def funcFiltroLP(senal, orden, fc, fs):
  [b,a] = signal.butter(orden, fc, btype = 'lowpass', analog = False, output = 'ba', fs = fs)
  senalLP = signal.filtfilt(b, a, senal)
  return senalLP

# Función para filtrar la señal con un filtro pasa altos
def funcFiltroHP(senal, orden, fc, fs):
  [b,a] = signal.butter(orden, fc, btype = 'highpass', analog = False, output = 'ba', fs = fs)
  senalHP = signal.filtfilt(b, a, senal)
  return senalHP

# Función para detectar picos 
def funcDetectarPicos(senal_ppg_filtrada):
  picos = signal.find_peaks(senal_ppg_filtrada)[0]
  return picos

def funEliminarPicosSubida(senal_ppg, locs_peaks_ppg):
  locs_peaks_ppg_correctos = []
  locs_peaks_ppg_correctos.append(locs_peaks_ppg[0])
  for loc in range(1, len(locs_peaks_ppg)):
    if (locs_peaks_ppg[loc]+10) < len(senal_ppg):
        if senal_ppg[locs_peaks_ppg[loc]+10] < senal_ppg[locs_peaks_ppg[loc]]:
            locs_peaks_ppg_correctos.append(locs_peaks_ppg[loc])
    else:
        locs_peaks_ppg_correctos.append(locs_peaks_ppg[loc])
  return np.array(locs_peaks_ppg_correctos)

def funcEliminarPicosOutliers(locs_peaks_ppg, low_rri, high_rri):
  locs_peaks_ppg_limpio = []
  locs_peaks_ppg_limpio.append(locs_peaks_ppg[0])
  for i in range(1,len(locs_peaks_ppg)):
    intervalo = locs_peaks_ppg[i]-locs_peaks_ppg_limpio[-1]
    if high_rri >= intervalo >= low_rri:
      locs_peaks_ppg_limpio.append(locs_peaks_ppg[i])
  return np.array(locs_peaks_ppg_limpio)

def funcEliminarDiastolicos(senal_ppg, locs_peaks_ppg):
  locs_peaks_ppg_sistolicos = []
  locs_peaks_ppg_sistolicos.append(locs_peaks_ppg[0])
  for loc in range(1, len(locs_peaks_ppg)):
    if senal_ppg[locs_peaks_ppg[loc]-10] < senal_ppg[locs_peaks_ppg[loc]]:
      locs_peaks_ppg_sistolicos.append(locs_peaks_ppg[loc])
  return np.array(locs_peaks_ppg_sistolicos)

# Función para detectar onsets
def funcDetectarOnsets(senal_ppg, locs_peaks_ppg):
  locs_onsets_ppg = []
  for i in range(len(locs_peaks_ppg)-1):
    locs_onsets_ppg.append(np.argmin(senal_ppg[locs_peaks_ppg[i]:locs_peaks_ppg[i+1]])+locs_peaks_ppg[i])
  return np.array(locs_onsets_ppg)

# Función para calcular el parámetro HBI como la diferencia temporal entre picos consecutivos
def funcHBI(locs_peaks_ppg_limpios):
  HBI = [j-i for i,j in zip(locs_peaks_ppg_limpios[0:-1], locs_peaks_ppg_limpios[1:])]
  return HBI

# Función para calcular el parámetro PPGA como la diferencia de amplitud entre onset y pico consecutivos 
def funcPPGA(senal_ppg, locs_peaks, locs_onsets):
  if locs_onsets[0]>locs_peaks[0]:
    locs_peaks = locs_peaks[1:]
  PPGA = [int(senal_ppg[peak]-senal_ppg[onset]) for peak,onset in zip(locs_peaks, locs_onsets)]
  return PPGA


############################################################################################################################################

# CAMBIAR CSV DATA
nombre_csv = 'DATA_LULI1.csv'

directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_csv_senal_ppg = nombre_csv
path_csv_ppg = os.path.join(directorio_actual, nombre_csv_senal_ppg)
senal_ppg = funcLevantarSenal(path_csv_ppg)

# CAMBIAR MUESTRAS
senal_ppg = senal_ppg[:-50]

fs = 50

senal_ppg_HP = funcFiltroHP(senal = senal_ppg, orden = 3, fc = 0.85, fs = fs)
senal_ppg_LP = funcFiltroLP(senal = senal_ppg_HP, orden = 3, fc = 5, fs = fs)

locs_peaks_ppg = funcDetectarPicos(senal_ppg_LP)
locs_peaks_ppg_no_subida = funEliminarPicosSubida(senal_ppg_LP, locs_peaks_ppg)
locs_peaks_ppg_sistolicos = funcEliminarDiastolicos(senal_ppg_LP, locs_peaks_ppg_no_subida)
locs_peaks_ppg_sin_outliers = funcEliminarPicosOutliers(locs_peaks_ppg_sistolicos, low_rri = 15, high_rri = 90)
locs_onsets_ppg = funcDetectarOnsets(senal_ppg_LP, locs_peaks_ppg_sin_outliers)

PPGA = funcPPGA(senal_ppg_LP, locs_peaks_ppg_sin_outliers, locs_onsets_ppg)
HBI = funcHBI(locs_peaks_ppg_sin_outliers)

# Excel Curva Normalizacion
directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_normalizacion = 'Curva Normalizacion.xlsx'
path_normalizacion = os.path.join(directorio_actual, nombre_excel_normalizacion)
excel_normalizacion = load_workbook(path_normalizacion)
hoja_excel_normalizacion = excel_normalizacion['Hoja1']

for i in range(len(HBI)):
    nueva_medicion = [HBI[i], PPGA[i], nombre_csv]
    hoja_excel_normalizacion.append(nueva_medicion)

excel_normalizacion.save(path_normalizacion)


# Excel Desvios 
directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_desvios = 'Desvios.xlsx'
path_desvios = os.path.join(directorio_actual, nombre_excel_desvios)
excel_desvios = load_workbook(path_desvios)
hoja_excel_desvios = excel_desvios['Hoja1']

desvio_HBI = np.std(HBI)
desvio_PPGA = np.std(PPGA)

desvios = [desvio_HBI, desvio_PPGA, nombre_csv]

hoja_excel_desvios.append(nueva_medicion)

excel_desvios.save(path_desvios)
