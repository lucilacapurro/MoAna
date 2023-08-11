import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import signal
import os
from openpyxl import load_workbook

################################################################################################################################
# FUNCIONES:

'''
# Función para obtener el histograma 
def funcHistograma(parametro, tipo_parametro):
  if tipo_parametro == "HBI":
    min = 20
    max = 200
  elif tipo_parametro == "PPGA":
    min = 0
    max = 8000
  bins = max-min
  vector_parametro = np.linspace(min, max, bins+1, dtype = int)
  vector_hist = np.zeros(bins+1)
  for i in range(len(parametro)):
    vector_hist[int(parametro[i])-min] += 1
  vector_hist_norm = vector_hist/np.sum(vector_hist)
  plt.figure(figsize=(20,5))
  plt.title('Histograma', fontsize=15)
  plt.bar(x = vector_parametro, height = vector_hist_norm, width=0.8, bottom=None, align='center')
  plt.show()
  TF_histograma = [[int(param), param_norm] for param, param_norm in zip(vector_parametro, vector_hist_norm)]
  return TF_histograma

# Función para obtener el histograma acumulado
def funcHistogramaAcumulado(TF_histograma):
  min = TF_histograma[0][0]
  max = TF_histograma[-1][0]
  bins = len(TF_histograma)
  vector_parametro = np.linspace(min, max, bins, dtype = int)
  vector_hist_acum = np.zeros(len(TF_histograma))
  vector_hist_acum[0] = TF_histograma[0][1]
  for i in range(1,len(TF_histograma)):
    vector_hist_acum[i] = vector_hist_acum[i-1] + TF_histograma[i][1]
  plt.figure(figsize=(10,5))
  plt.title('Histograma Acumulado', fontsize=15)
  plt.bar(x = vector_parametro, height = vector_hist_acum, width=0.8, bottom=None, align='center')
  plt.show()
  TF_histograma_acum = [[int(param), round(param_norm, 7)] for param, param_norm in zip(vector_parametro, vector_hist_acum)]
  return TF_histograma_acum
'''

# Función para normalizar el parámetro de entrada según su funcion de transformación
def funcNormalizarParametro(TF_normalizacion, parametro):
  parametro = int(parametro)
  min = TF_normalizacion[0][0]
  parametro_norm = TF_normalizacion[parametro-min][1]
  return parametro_norm

# Función para hacer el promedio de una lista de valores 
def funcPromedio(vector_valores):
  prom = np.mean(vector_valores)
  return int(prom)

# Función para calcular el valor del SPI instantáneo a partir de un valor de HBI y PPGA normalizado
def funcSPIi(PPGAi_norm, HBIi_norm):
  SPIi = 100 - (0.67*PPGAi_norm+0.33*HBIi_norm)*100
  return int(SPIi)

################################################################################################################################
# GENERA LAS FUNCIONES DE TRANSFORMACIÓN:

# Obtiene los valores de HBI y PPGA de entrada del excel de valores HBI y PPGA de registros hechos:
directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_normalizacion = 'Curva Normalizacion.xlsx'
path_normalizacion = os.path.join(directorio_actual, nombre_excel_normalizacion)
excel_normalizacion = pd.read_excel(path_normalizacion)

# Obtiene las listas de los valores de los parámetros de entrada 
lista_poblacional_HBI = excel_normalizacion['HBI'].dropna()
lista_poblacional_PPGA = excel_normalizacion['PPGA'].dropna()

# Calcula los mínimo y máximos para hacer el eje x de entrada 
# HBI
min_poblacional_HBI = np.min(lista_poblacional_HBI)
max_poblacional_HBI = np.max(lista_poblacional_HBI)
x_HBI = np.linspace(min_poblacional_HBI, max_poblacional_HBI, max_poblacional_HBI-min_poblacional_HBI+1)
x_HBI = x_HBI.tolist()
# PPGA 
min_poblacional_PPGA = np.min(lista_poblacional_PPGA)
max_poblacional_PPGA = np.max(lista_poblacional_PPGA)
x_PPGA = np.linspace(min_poblacional_PPGA, max_poblacional_PPGA, max_poblacional_PPGA-min_poblacional_PPGA+1)
x_PPGA = x_PPGA.tolist()

# Obtiene los valores de HBInorm y PPGAnorm de salida obtenidos de la normalización gaussiana:
directorio_actual = os.path.abspath(os.path.dirname(__file__))
nombre_excel_normalizacion = 'Curva Normalizacion Gaussiana.xlsx'
path_normalizacion = os.path.join(directorio_actual, nombre_excel_normalizacion)
excel_normalizacion = pd.read_excel(path_normalizacion)

# Obtiene las listas de los valores de los parámetros normalizados 
HBI_curva_excel = np.array(excel_normalizacion['HBI'].dropna())
PPGA_curva_excel = np.array(excel_normalizacion['PPGA'].dropna())

# Hace los acumulados (valores de salida y correspondientes a los valores de entrada x)
y_HBI = np.cumsum(HBI_curva_excel)
y_HBI = y_HBI.tolist()
y_PPGA = np.cumsum(PPGA_curva_excel)
y_PPGA = y_PPGA.tolist()
# Genera las transformaciones de mapeo en formato de data frame
df_TF_HBI = pd.DataFrame({'x HBI': x_HBI, 'y HBI': y_HBI})
df_TF_PPGA = pd.DataFrame({'x PPGA': x_PPGA, 'y PPGA': y_PPGA})

# Guarda las TF en los excels:
# HBI
nombre_excel_TF_HBI = 'TF_HBI_Gaussiana.xlsx'
path_excel_TF_HBI = os.path.join(directorio_actual, nombre_excel_TF_HBI)
excel_TF_HBI = load_workbook(path_excel_TF_HBI)
df_TF_HBI.to_excel(path_excel_TF_HBI, index=False, header=False)
# PPGA
nombre_excel_TF_PPGA = 'TF_PPGA_Gaussiana.xlsx'
path_excel_TF_PPGA = os.path.join(directorio_actual, nombre_excel_TF_PPGA)
excel_TF_PPGA = load_workbook(path_excel_TF_PPGA)
df_TF_PPGA.to_excel(path_excel_TF_PPGA, index=False, header=False)


############################################################################################
# GRÁFICOS:

plt.figure(figsize=(10,5))
plt.title('Histograma Acumulado HBI', fontsize=15)
plt.bar(x = x_HBI, height = y_HBI, width=0.8, bottom=None, align='center')
plt.show()

plt.figure(figsize=(10,5))
plt.title('Histograma Acumulado PPGA', fontsize=15)
plt.bar(x = x_PPGA, height = y_PPGA, width=0.8, bottom=None, align='center')
plt.show()


################################################################################################################################
# PRUEBAS:
'''
#PPGA = [585, 607, 619, 540, 604, 588, 583, 546, 587, 573, 552]
#HBI = [102, 99, 98, 93, 92, 93, 95, 93, 95, 98, 100]

#PPGA = [724, 1064, 945, 771, 895, 1359, 1036, 855, 1180, 1148, 767, 1014, 1123, 1419, 677, 1316, 712, 1309, 1125, 992, 1122, 450, 678, 557, 390, 699, 758, 390, 1091, 1176, 1619, 937, 1133, 704, 970, 310, 1048, 754, 536, 542, 831, 786, 866, 698, 747, 1115, 792, 966, 1357, 1109, 817, 934, 1462, 1011, 1017, 1124, 1332, 1405, 1150, 818, 1115, 854, 774, 706, 713, 602, 829, 766, 537, 469, 951, 687, 439, 711, 456, 533, 644, 446, 444, 450, 418, 543, 607, 740, 957, 778, 631, 873, 1115, 1067, 775, 884, 1416, 1403, 974, 936, 1243, 1408, 1017, 957, 1336, 1324, 1016, 894, 1070, 803, 698, 897, 1207, 993, 932, 1015, 1306, 1511, 1139, 962, 1178, 1467, 1287, 1022, 1078, 1370, 1290, 1225, 952, 905, 1104, 1292, 1188, 1130, 1609, 1583, 1221, 1307, 1809, 1484, 1511, 1376, 1764, 786, 948, 1745, 954, 1024, 984, 819, 911, 860, 881, 939, 809, 792, 987, 854, 755, 913, 1195, 1045, 923, 1179, 1114, 964, 1105, 1217, 983, 1105, 1331, 1021, 947, 1125, 1148, 955, 1146, 1394, 1140, 1101, 1345, 1181, 1017, 1164, 1285, 1058, 1027, 1196, 1007, 888, 1008, 819]
#HBI = [77, 85, 86, 78, 86, 89, 86, 77, 87, 88, 83, 80, 84, 130, 41, 75, 67, 74, 78, 80, 99, 53, 84, 78, 78, 77, 76, 73, 80, 79, 95, 75, 81, 87, 77, 83, 73, 137, 65, 60, 72, 70, 85, 72, 78, 92, 80, 80, 98, 94, 83, 81, 90, 85, 84, 80, 83, 86, 88, 81, 84, 75, 74, 72, 69, 70, 75, 75, 77, 69, 103, 42, 68, 71, 63, 67, 71, 65, 68, 68, 66, 67, 70, 67, 72, 77, 70, 75, 78, 84, 80, 77, 81, 85, 85, 75, 84, 89, 83, 76, 86, 90, 79, 77, 92, 87, 79, 84, 92, 86, 81, 80, 90, 92, 89, 79, 81, 84, 87, 74, 78, 81, 87, 87, 82, 80, 83, 86, 86, 79, 83, 84, 79, 77, 87, 85, 78, 80, 88, 70, 72, 84, 66, 73, 68, 69, 71, 69, 76, 76, 82, 75, 81, 84, 76, 76, 81, 81, 79, 82, 89, 79, 80, 82, 82, 79, 86, 85, 81, 83, 91, 84, 82, 89, 81, 78, 84, 87, 78, 82, 91, 84, 83, 99, 91, 81, 91, 88]

# REPOSO LULI
#PPGA = [3373, 3620, 3640, 3248, 3387, 3446, 3569, 2995, 3182, 3332, 3306, 2955, 2963, 3362, 3407, 3394, 2860, 3220, 3254, 3312, 2882, 3259, 3379, 3402, 2844, 3079, 3276, 3264, 2924, 2978, 3223, 2945, 2853, 2791, 3017, 3080, 2953, 2665, 3000, 3051, 2948, 2689, 3085, 3149, 3015, 2704, 2877, 2994, 3083, 2693, 2664, 3047, 3133, 2784, 2704, 3118, 3040, 2797, 2661, 2917, 3191, 2898, 2662, 3013, 3078, 2871, 2661, 2943, 2972, 2752, 2687, 3075, 3133, 2866, 2992, 3166, 2912, 2655, 3061, 2930, 2401, 2619]
#HBI = [89, 92, 95, 92, 88, 89, 93, 86, 84, 86, 89, 87, 83, 86, 91, 93, 85, 84, 87, 91, 86, 86, 91, 94, 86, 83, 88, 90, 87, 82, 85, 86, 86, 81, 84, 89, 89, 82, 84, 88, 90, 82, 86, 89, 89, 82, 81, 87, 91, 87, 82, 90, 96, 90, 83, 89, 93, 90, 83, 88, 97, 93, 84, 88, 92, 92, 81, 87, 90, 90, 83, 92, 95, 92, 87, 93, 97, 86, 92, 91, 87, 81]
# VECTOR SPI = [14, 11, 9, 12, 15, 14, 10, 18, 19, 16, 14, 17, 21, 16, 12, 11, 20, 19, 16, 13, 19, 17, 13, 10, 19, 20, 15, 14, 18, 21, 18, 18, 19, 24, 20, 15, 16, 23, 20, 16, 15, 23, 17, 15, 15, 23, 23, 17, 14, 19, 24, 15, 10, 16, 22, 15, 13, 16, 23, 17, 9, 13, 22, 16, 13, 14, 24, 17, 15, 17, 23, 13, 10, 14, 17, 12, 10, 20, 13, 15, 21, 25]

# DOLOR LULI
#PPGA = [3142, 3127, 3115, 2861, 3025, 2876, 1952, 1443, 1244, 1233, 1139, 1045, 1129, 1219, 1025, 800, 754, 2124, 1516, 1753, 1577, 1733, 1884, 1929, 2149, 2012, 2110, 2250, 3057, 2275, 2106, 2269, 2150, 2037, 2044, 2189, 2253, 2104, 1997, 2190, 2229, 2091, 1998, 2209, 2068, 1997, 1610, 2035, 1979, 1940, 1723, 1885, 1559, 1473, 1466, 1786, 1581, 3861, 1206, 1978, 1851, 2279, 1597, 2384, 2283, 2187, 2483]
#HBI = [94, 99, 93, 85, 89, 94, 91, 83, 81, 83, 79, 75, 79, 81, 79, 74, 74, 75, 76, 78, 76, 76, 76, 78, 80, 77, 77, 82, 84, 86, 77, 79, 78, 76, 75, 79, 85, 89, 81, 86, 92, 87, 80, 80, 81, 82, 77, 80, 84, 83, 77, 80, 81, 79, 75, 76, 76, 79, 81, 86, 85, 88, 80, 88, 90, 94, 84]
# VECTOR SPI = [11, 8, 12, 20, 15, 12, 25, 43, 49, 48, 53, 60, 54, 50, 56, 68, 69, 36, 49, 41, 47, 43, 40, 37, 30, 36, 34, 27, 19, 23, 34, 30, 32, 36, 37, 31, 24, 23, 32, 24, 19, 25, 33, 30, 31, 32, 45, 32, 30, 32, 42, 36, 42, 46, 51, 42, 47, 21, 50, 29, 32, 22, 42, 21, 20, 18, 23]

# REPOSO ZAKI
#PPGA = [1519, 1703, 1958, 2671, 2805, 2902, 2815, 2810, 3180, 3040, 2786, 2924, 3109, 3154, 3139, 2876, 3280, 3363, 3290, 3269, 2934, 3071, 3305, 3325, 3282, 2924, 3419, 3457, 3531, 3438, 3012, 2963, 3244, 3268, 3309, 3300, 2876, 3092, 3407, 3555, 3476, 3203, 2954, 3427, 3417, 3312, 2983, 2952, 3012, 2763, 2620, 2131, 2354, 2809, 2550, 2081, 1562, 1721, 2071, 2275, 2208, 1951, 2284, 2587, 2530, 2204, 2167, 2750, 2825, 2718, 2492, 2335, 2787, 3051, 2996, 2661, 2430, 2898, 3113, 2581, 1995, 1842, 2336, 2479, 2442, 2178, 2510, 2796, 2882, 2749, 2580, 3150, 3014, 2573, 2156, 2840, 2843, 2761, 2682, 3170, 3313, 3148, 2808, 3187, 3329, 3178, 2877, 2660, 3248, 3320, 3314, 2947, 3067, 3515, 3287, 3006, 2936, 2808, 2683, 3127, 3362, 3468, 3509, 3281, 3854, 3742, 3552, 3309, 3629, 3515, 3426, 3335, 3079, 3552, 3477, 3167, 2744, 3149, 3104, 2860, 2409, 3170]
#HBI = [95, 92, 83, 85, 85, 86, 87, 81, 83, 84, 87, 83, 85, 86, 86, 80, 80, 81, 83, 85, 82, 77, 80, 83, 86, 80, 78, 80, 83, 82, 78, 73, 73, 76, 79, 82, 77, 73, 75, 80, 82, 83, 77, 78, 79, 81, 79, 73, 75, 80, 84, 79, 75, 78, 82, 85, 78, 74, 75, 76, 78, 76, 71, 76, 81, 80, 75, 76, 80, 83, 82, 75, 74, 76, 81, 82, 76, 75, 79, 83, 82, 75, 78, 80, 82, 79, 76, 79, 81, 81, 75, 75, 79, 85, 81, 80, 84, 85, 76, 77, 83, 85, 77, 76, 78, 84, 87, 79, 80, 84, 86, 82, 75, 78, 81, 73, 72, 72, 71, 75, 81, 88, 90, 88, 86, 87, 87, 79, 80, 82, 82, 83, 79, 82, 84, 84, 80, 77, 80, 82, 70, 71]
# VECTOR SPI = [32, 30, 32, 21, 20, 19, 18, 23, 20, 20, 18, 21, 18, 17, 17, 24, 22, 21, 19, 17, 22, 26, 22, 19, 17, 24, 23, 21, 18, 19, 25, 30, 29, 26, 23, 20, 27, 30, 27, 21, 19, 19, 26, 23, 22, 21, 24, 31, 28, 25, 22, 32, 33, 26, 24, 27, 45, 46, 37, 33, 31, 38, 37, 30, 25, 30, 35, 29, 24, 22, 25, 33, 31, 27, 22, 24, 31, 29, 24, 23, 32, 42, 30, 27, 25, 31, 30, 26, 23, 24, 31, 28, 24, 22, 29, 24, 21, 20, 29, 25, 19, 18, 27, 26, 24, 19, 18, 26, 22, 18, 16, 22, 28, 23, 21, 30, 31, 32, 34, 28, 21, 14, 13, 15, 15, 14, 15, 23, 21, 19, 19, 19, 24, 19, 18, 19, 25, 25, 23, 22, 36, 31]

# DOLOR ZAKI
PPGA = [1954, 1967, 1843, 2030, 2138, 1717, 2030, 1972, 2057, 2415, 2382, 2339, 1989, 1902, 1991, 1947, 2114, 2388, 2523, 2464, 2203, 2490, 2808, 2669, 2218, 2219, 1962, 1680, 1983, 2082, 1998, 2555, 2599, 2273, 2578, 2361, 2181, 2347, 2575, 2208, 1587, 1597, 1445, 1392, 1413, 1469, 1257, 1814, 1924, 1832, 1809, 1763, 1642, 1370, 1585, 1485, 1591, 1704, 1797, 1647, 1475, 1886, 1899, 1846, 1671, 2275, 2240, 2001, 2451, 2374, 2361, 2441, 2450, 2066, 1768, 1617, 1414, 1282, 1157, 1656, 2046, 2060, 2211, 2230, 2467, 2196, 2516, 2607, 2642, 2362, 2555, 2813, 2779, 2535, 2910, 2822, 2373, 2250, 2601, 2925, 2927, 2710, 2464, 2661, 2435, 2584, 2148, 1451, 1314, 1613]
HBI = [81, 81, 80, 77, 79, 78, 79, 82, 81, 84, 83, 82, 79, 78, 80, 78, 74, 76, 79, 84, 79, 74, 77, 80, 75, 80, 82, 82, 78, 82, 80, 82, 84, 80, 79, 80, 81, 77, 78, 78, 75, 77, 83, 81, 86, 89, 82, 81, 80, 77, 73, 75, 77, 75, 77, 79, 81, 77, 78, 79, 75, 75, 74, 78, 75, 78, 79, 77, 76, 77, 76, 74, 73, 74, 73, 78, 81, 79, 75, 74, 76, 78, 77, 78, 79, 78, 75, 75, 78, 77, 72, 74, 75, 75, 73, 76, 77, 75, 72, 74, 77, 80, 79, 81, 75, 73, 73, 73, 71, 74]
# VECTOR SPI = [33, 33, 37, 36, 31, 41, 34, 32, 31, 24, 25, 26, 35, 37, 34, 36, 37, 32, 27, 23, 31, 33, 27, 25, 35, 29, 32, 38, 36, 30, 33, 24, 22, 29, 27, 28, 29, 31, 28, 31, 48, 45, 43, 46, 42, 38, 48, 36, 35, 40, 44, 44, 44, 53, 46, 46, 41, 43, 39, 42, 51, 41, 42, 38, 46, 30, 30, 36, 31, 30, 32, 33, 34, 38, 45, 44, 46, 51, 57, 47, 36, 34, 32, 31, 28, 31, 32, 31, 28, 31, 34, 31, 30, 32, 31, 28, 30, 34, 34, 30, 26, 25, 28, 24, 32, 33, 38, 53, 57, 48]

# Calcula los HBI y PPGA promedio de los vectores
HBI_prom = funcPromedio(HBI)
PPGA_prom = funcPromedio(PPGA)
# Normaliza los parámetros 
HBI_norm = funcNormalizarParametro(TF_HBI, int(HBI))
PPGA_norm = funcNormalizarParametro(TF_PPGA, int(PPGA))

# Crea un vector con los valores de SPI calculados
vectorSPI = []
for i in range(len(PPGA)): 
  PPGA_norm = funcNormalizarParametro(TF_PPGA, int(PPGA[i]))
  HBI_norm = funcNormalizarParametro(TF_HBI, int(HBI[i]))
  SPI = funcSPIi(PPGA_norm, HBI_norm)
  vectorSPI.append(SPI)
print(vectorSPI)

# Crear una lista de índices para el eje x
x = range(len(vectorSPI))
# Crea el gráfico 
plt.figure()
plt.plot(x, vectorSPI)
plt.title("Gráfico de evolución del índice SPI")
plt.xlabel("Tiempo")
plt.ylabel("SPI")
plt.ylim(0,100)
plt.show()
'''